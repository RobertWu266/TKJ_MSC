import openpyxl
from tkinter import Tk, filedialog, messagebox, simpledialog, Toplevel, ttk

def create_excel(name_list, value_input):
    root = Tk()
    root.withdraw()
    
    # Check if the user wants to create a new excel file
    is_new_file = messagebox.askyesno("Create Excel File", "Do you want to create a new excel file?")
    
    if is_new_file:
        # Create new excel file and set the file name
        file_path = filedialog.asksaveasfilename(initialdir="/home/wu/PycharmProjects/TKJ_MSC",defaultextension=".xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        sheet_name = simpledialog.askstring("Enter the worksheet name", "Please enter the worksheet name:")
        ws.title=sheet_name
    else:
        # Open an existing excel file
        file_path = filedialog.askopenfilename(initialdir='/',defaultextension=".xlsx")
        wb = openpyxl.load_workbook(file_path)
        
        # Check if the user wants to create a new worksheet
        is_new_sheet = messagebox.askyesno("Create Worksheet", "Do you want to create a new worksheet?")
    
        if is_new_sheet:
            # Create new worksheet and set the sheet name
            sheet_name = simpledialog.askstring("Enter the worksheet name", "Please enter the worksheet name:")
            ws = wb.create_sheet(title=sheet_name)
        else:
            # Open an existing worksheet
            sheets = wb.sheetnames
            top = Toplevel()
            top.title("Select the existing worksheet")
            cb = ttk.Combobox(top, values=sheets, state='readonly')
            button = ttk.Button(top, text="OK", command=top.destroy)
            cb.pack()
            cb.current(0)
            try:
                button.pack()
            except:
                pass
            
            while True:
                try:
                    if top.state() == 'normal':
                        top.update()
                        sheet_name = cb.get()
                        cb.wait_variable(var=button)
                        continue
                    else:
                        break
                except:
                    break
                
            
            try:
                top.mainloop()
            except:
                pass
            
            ws = wb[sheet_name]
    
    #print('s')
    #Write the data of name_list and value_input
    for i in range(len(name_list)):
        ws.cell(row=i+2, column=1).value = name_list[i]
        ws.cell(row=1, column=i+2).value = name_list[i]
        for j in range(len(value_input[i])):
            ws.cell(row=i+2, column=j+2).value = value_input[i][j]
    
    # Save the workbook
    wb.save(file_path)
    
    messagebox.showinfo("Excel File Created", "The excel file has been created and saved at " + file_path)


